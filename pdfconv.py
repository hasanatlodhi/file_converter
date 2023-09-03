
import aspose.pdf as ap
import openpyxl
def generate_excel_from_pdf(input_pdf,output_excel):

    # Open PDF document
    document = ap.Document(input_pdf)

    save_option = ap.ExcelSaveOptions()
    # save_option.format = ap.ExcelSaveOptions.ExcelFormat.XLSX
    # Save the file into MS Excel format
    document.save(output_excel, save_option)
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(output_excel)

    # Choose the specific worksheet where the text is located

    for sheetname in  workbook.sheetnames:
        worksheet = workbook[sheetname]
    # Iterate through the cells and remove the desired text
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                try:
                    if cell.value is not None and 'Evaluation Only' in cell.value:
                        cell.value = cell.value.replace('Evaluation Only. Created with Aspose.PDF. Copyright 2002-2023 Aspose Pty Ltd.', '')
                except:
                    pass
    # Save the modified workbook
    workbook.save(output_excel)

# generate_excel_from_pdf('scccs.pdf','output.xlsx)


# Read a PDF File


